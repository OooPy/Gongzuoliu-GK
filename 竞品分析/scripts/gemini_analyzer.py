# -*- coding: utf-8 -*-
"""
Gemini AI 竞品分析模块
医药保健品行业大师级电商竞品分析引擎

分析维度：
1. 5张主图卖点与视觉策略
2. 详情页表达逻辑思路
3. 规格设计分析
4. 活动与促销策略
5. 评价数量与消费者关注点
6. 近30日销售数据分析
7. SKU销售数据分析
8. 搜索词与转化分析
9. 流量来源分析
"""

import os
import json
import time
import random
import google.generativeai as genai
from pathlib import Path
from PIL import Image
from dotenv import load_dotenv

# 加载环境变量
script_dir = Path(__file__).parent.resolve()
env_path = script_dir.parent / '.env'
load_dotenv(env_path)

# Gemini 配置
API_KEY = os.getenv("GOOGLE_API_KEY")
MODEL_NAME = "gemini-2.5-flash"

_initialized = False

# ============================================================
# 行业大师级专家人设
# ============================================================

EXPERT_PERSONA = """你是一位在医药保健品行业深耕15年以上的电商运营大师级专家。

【你的专业背景】
- 曾操盘过多个知名保健品品牌（东阿阿胶、同仁堂、汤臣倍健等级别）的天猫/京东旗舰店，年GMV过亿
- 深度理解保健品/OTC药品行业的电商特性：合规红线、功效表达技巧、信任构建路径
- 精通阿胶、人参、燕窝、益生菌、维生素等滋补保健品赛道的竞争格局与消费趋势
- 对保健品消费者心理有深刻洞察：送礼 vs 自用场景差异、功效焦虑与信任博弈、价格敏感度分层
- 熟悉各电商平台保健品类目的流量分发逻辑、搜索排名机制、活动报名规则

【分析原则】
- 基于数据和截图做客观分析，不凭空臆造数据
- 结合保健品行业特性给出专业解读，引用行业常识和经验
- 分析要有深度但保持实战导向，给出可落地的见解
- 如果某些数据不完整或无法获取，明确指出并给出基于行业经验的合理推断
- 用电商运营的专业语言表达，体现行业老手的判断力"""


# ============================================================
# 基础工具函数
# ============================================================

def init_gemini():
    """初始化 Gemini API（仅首次）"""
    global _initialized
    if _initialized:
        return
    if not API_KEY:
        raise ValueError("请在 .env 文件中设置 GOOGLE_API_KEY")
    genai.configure(api_key=API_KEY)
    _initialized = True


def _call_gemini_vision(image_paths: list, prompt: str, max_retries: int = 5) -> str:
    """调用 Gemini 视觉 API（支持多图）"""
    init_gemini()

    contents = [prompt]
    for img_path in image_paths:
        contents.append(Image.open(img_path))

    model = genai.GenerativeModel(MODEL_NAME)

    for attempt in range(max_retries):
        try:
            response = model.generate_content(contents)
            return response.text.strip()
        except Exception as e:
            if attempt < max_retries - 1:
                err_str = str(e)
                if '429' in err_str or 'Resource exhausted' in err_str:
                    wait_time = (attempt + 1) * 15 + random.uniform(0, 3)
                    print(f"    ⚠ API 限频，{wait_time:.0f}秒后重试 ({attempt+1}/{max_retries}): 429 Rate Limited")
                else:
                    wait_time = (attempt + 1) * 3 + random.uniform(0, 1)
                    print(f"    ⚠ API 调用失败，{wait_time:.1f}秒后重试 ({attempt+1}/{max_retries}): {e}")
                time.sleep(wait_time)
            else:
                raise RuntimeError(f"API 调用失败（已重试{max_retries}次）: {e}")
    return ""


def _call_gemini_text(prompt: str, max_retries: int = 5) -> str:
    """调用 Gemini 纯文本 API"""
    init_gemini()
    model = genai.GenerativeModel(MODEL_NAME)

    for attempt in range(max_retries):
        try:
            response = model.generate_content(prompt)
            return response.text.strip()
        except Exception as e:
            if attempt < max_retries - 1:
                err_str = str(e)
                if '429' in err_str or 'Resource exhausted' in err_str:
                    wait_time = (attempt + 1) * 15 + random.uniform(0, 3)
                    print(f"    ⚠ API 限频，{wait_time:.0f}秒后重试 ({attempt+1}/{max_retries}): 429 Rate Limited")
                else:
                    wait_time = (attempt + 1) * 3 + random.uniform(0, 1)
                    print(f"    ⚠ API 调用失败，{wait_time:.1f}秒后重试 ({attempt+1}/{max_retries}): {e}")
                time.sleep(wait_time)
            else:
                raise RuntimeError(f"API 调用失败（已重试{max_retries}次）: {e}")
    return ""


def _parse_json_response(text: str) -> dict:
    """从 Gemini 返回文本中提取 JSON"""
    import re

    # 尝试直接解析
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # 尝试提取 ```json ... ```
    json_match = re.search(r'```(?:json)?\s*\n?(.*?)\n?\s*```', text, re.DOTALL)
    if json_match:
        try:
            return json.loads(json_match.group(1))
        except json.JSONDecodeError:
            pass

    # 尝试找到 { ... }
    brace_match = re.search(r'\{.*\}', text, re.DOTALL)
    if brace_match:
        try:
            return json.loads(brace_match.group())
        except json.JSONDecodeError:
            pass

    return {"raw_text": text, "_parse_failed": True}


# ============================================================
# 数据文件读取（Excel / PDF / CSV，全部本地处理，不消耗 API）
# ============================================================

def read_excel_data(file_path: str, max_chars: int = 50000) -> str:
    """
    读取 Excel 文件并转换为结构化文本

    Args:
        file_path: Excel 文件路径
        max_chars: 最大字符数限制（避免 token 溢出）

    Returns:
        str: 格式化的文本数据
    """
    try:
        import openpyxl
    except ImportError:
        print("    ⚠ openpyxl 未安装，跳过 Excel 数据读取")
        return ""

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"    ⚠ 读取 Excel 失败: {e}")
        return ""

    sections = []
    total_chars = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text = []

        for row in ws.iter_rows(values_only=True):
            values = []
            for v in row:
                if v is None:
                    values.append("")
                elif isinstance(v, float):
                    # 避免过长的小数
                    if v == int(v):
                        values.append(str(int(v)))
                    else:
                        values.append(f"{v:.2f}")
                else:
                    values.append(str(v))
            if any(v.strip() for v in values):
                rows_text.append(" | ".join(values))

        if rows_text:
            section = f"\n{'='*40}\n【{sheet_name}】\n{'='*40}\n" + "\n".join(rows_text)
            total_chars += len(section)
            if total_chars > max_chars:
                section += f"\n... (数据过多，已截断，共 {len(rows_text)} 行)"
                sections.append(section)
                break
            sections.append(section)

    wb.close()
    return "\n".join(sections)


def read_all_data_files(data_files: dict) -> tuple:
    """
    读取所有数据文件并合并为文本

    Args:
        data_files: {"xlsx": [...], "pdf": [...], "csv": [...]}

    Returns:
        tuple: (合并后的文本数据, 图片型PDF转出的图片路径列表)
    """
    all_text = []

    # 读取 Excel 文件
    for xlsx_path in data_files.get("xlsx", []):
        filename = Path(xlsx_path).name
        print(f"    📊 读取数据文件: {filename}")
        text = read_excel_data(xlsx_path)
        if text:
            all_text.append(f"\n{'#'*50}\n数据来源: {filename}\n{'#'*50}\n{text}")

    # 读取 CSV 文件
    for csv_path in data_files.get("csv", []):
        filename = Path(csv_path).name
        print(f"    📊 读取数据文件: {filename}")
        try:
            with open(csv_path, 'r', encoding='utf-8') as f:
                text = f.read(10000)
            all_text.append(f"\n数据来源: {filename}\n{text}")
        except Exception:
            try:
                with open(csv_path, 'r', encoding='gbk') as f:
                    text = f.read(10000)
                all_text.append(f"\n数据来源: {filename}\n{text}")
            except Exception as e:
                print(f"    ⚠ 读取 CSV 失败: {e}")

    # 读取 PDF 文件（自动判断文本型/图片型）
    pdf_image_paths = []  # 收集图片型PDF转出的图片
    for pdf_path in data_files.get("pdf", []):
        filename = Path(pdf_path).name
        print(f"    📄 读取PDF文件: {filename}")
        pdf_result = read_pdf_data(pdf_path)
        if pdf_result["type"] == "text":
            all_text.append(f"\n{'#'*50}\n数据来源: {filename}\n{'#'*50}\n{pdf_result['text']}")
        elif pdf_result["type"] == "image":
            pdf_image_paths.extend(pdf_result["image_paths"])

    return "\n".join(all_text), pdf_image_paths


def read_pdf_data(file_path: str, max_chars: int = 20000) -> dict:
    """
    读取 PDF 文件，自动判断是文本型还是图片型 PDF

    Args:
        file_path: PDF 文件路径
        max_chars: 最大字符数限制

    Returns:
        dict: {
            "type": "text" | "image" | "empty",
            "text": 提取的文本（文本型PDF）,
            "image_paths": [转出的图片路径列表]（图片型PDF）,
            "page_count": 页数
        }
    """
    result = {"type": "empty", "text": "", "image_paths": [], "page_count": 0}

    # ====== Phase 1: 尝试文本提取 ======
    text_extracted = ""

    # 尝试 PyPDF2
    try:
        import PyPDF2
        text_parts = []
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            result["page_count"] = len(reader.pages)
            for i, page in enumerate(reader.pages):
                page_text = page.extract_text()
                if page_text and page_text.strip():
                    text_parts.append(f"--- 第{i+1}页 ---\n{page_text.strip()}")
        if text_parts:
            text_extracted = "\n\n".join(text_parts)
    except ImportError:
        pass
    except Exception as e:
        print(f"    ⚠ PyPDF2 读取异常: {e}")

    # 如果 PyPDF2 没提取到，尝试 pdfplumber
    if not text_extracted:
        try:
            import pdfplumber
            text_parts = []
            with pdfplumber.open(file_path) as pdf:
                result["page_count"] = len(pdf.pages)
                for i, page in enumerate(pdf.pages):
                    page_text = page.extract_text()
                    if page_text and page_text.strip():
                        text_parts.append(f"--- 第{i+1}页 ---\n{page_text.strip()}")
            if text_parts:
                text_extracted = "\n\n".join(text_parts)
        except ImportError:
            pass
        except Exception as e:
            print(f"    ⚠ pdfplumber 读取异常: {e}")

    # 如果提取到了文本，返回文本型结果
    if text_extracted and len(text_extracted.strip()) > 50:
        if len(text_extracted) > max_chars:
            text_extracted = text_extracted[:max_chars] + f"\n... (PDF内容已截断，共 {result['page_count']} 页)"
        print(f"       ✓ 文本型PDF，已提取 {len(text_extracted)} 字符")
        result["type"] = "text"
        result["text"] = text_extracted
        return result

    # ====== Phase 2: 文本提取失败/过少，判定为图片型PDF，用 PyMuPDF 转图片 ======
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(file_path)
        result["page_count"] = len(doc)

        pdf_dir = Path(file_path).parent
        pdf_stem = Path(file_path).stem
        image_paths = []

        for i, page in enumerate(doc):
            # 2x 缩放确保图片清晰度
            mat = fitz.Matrix(2.0, 2.0)
            pix = page.get_pixmap(matrix=mat)

            # 保存为临时图片（以 _pdf_page_ 前缀标记，便于后续清理）
            img_filename = f"_pdf_page_{pdf_stem}_{i+1}.png"
            img_path = str(pdf_dir / img_filename)
            pix.save(img_path)
            image_paths.append(img_path)
            print(f"       ✓ PDF第{i+1}页 → 图片 ({pix.width}x{pix.height})")

        doc.close()

        if image_paths:
            print(f"       ✓ 图片型PDF，已转出 {len(image_paths)} 张图片，将用视觉API分析")
            result["type"] = "image"
            result["image_paths"] = image_paths
            return result

    except ImportError:
        print(f"    ⚠ PyMuPDF 未安装，无法处理图片型PDF（pip install PyMuPDF）")
    except Exception as e:
        print(f"    ⚠ PyMuPDF 转图片失败: {e}")

    print(f"    ⚠ PDF无法读取（既无文本也无法转图片）")
    return result


# ============================================================
# 维度一：5张主图卖点与视觉策略分析
# ============================================================

def analyze_main_images_strategy(image_paths: list, text_info: str = "") -> dict:
    """
    分析5张电商主图的卖点与视觉策略

    Returns:
        dict: 逐张主图分析 + 整体策略评价
    """
    extra = f"\n\n【补充背景信息】\n{text_info}" if text_info else ""
    n = len(image_paths)

    prompt = f"""{EXPERT_PERSONA}

现在请你对以下{n}张电商主图进行专业的竞品分析。

【分析任务 — 主图卖点与视觉策略】

请逐张分析每张主图（按传入顺序从第1张到第{n}张），对每张主图分析：

1. **主图角色定位**：这张图在5张主图体系中承担什么角色？
   - 常见角色：首图（引流转化）、卖点展示、场景化/人群定位、促销利益点、品牌背书/资质
2. **核心卖点提炼**：这张图主要在传达什么卖点？卖点是否清晰有力？
3. **文案策略解析**：文案用了什么表达技巧？
   - 保健品常用技巧：数据化功效暗示、原料溯源、权威背书、使用场景具象化、恐惧营销（亚健康焦虑）、利益量化
4. **视觉设计思路**：色调、构图、元素选择背后的策略意图
   - 在保健品品类中这种视觉风格代表什么定位？
5. **行业合规评估**：保健品广告法约束下，是否存在违规风险？（如暗示疗效、绝对化用语等）

分析完所有主图后，给出 **整体主图体系总评**：
- 5张主图的整体策略逻辑线是什么？信息传达的递进关系如何？
- 卖点排布是否合理？是否存在重复或遗漏？
- 在保健品赛道中，这套主图的竞争力如何？
- 最值得学习/借鉴的点是什么？
- 1-2条关键改进建议
{extra}

【输出格式】请严格输出 JSON：

```json
{{
    "product_name": "品牌名+产品名",
    "images_analysis": [
        {{
            "image_no": 1,
            "role": "主图角色定位描述",
            "selling_point": "核心卖点分析",
            "copy_strategy": "文案策略解析",
            "visual_design": "视觉设计思路",
            "compliance_note": "合规评估"
        }}
    ],
    "overall_strategy": "5张主图的整体策略逻辑线描述",
    "selling_points_layout": "卖点排布评价",
    "competitive_rating": "在保健品赛道中的竞争力评价",
    "best_practices": "最值得借鉴的点",
    "improvement_suggestions": ["改进建议1", "改进建议2"]
}}
```"""

    result_text = _call_gemini_vision(image_paths, prompt)
    return _parse_json_response(result_text)


# ============================================================
# 维度二：详情页表达逻辑思路
# ============================================================

def analyze_detail_page_logic(image_paths: list, text_info: str = "") -> dict:
    """
    分析详情页的内容逻辑、说服路径、表达策略

    Returns:
        dict: 详情页逻辑分析结果
    """
    extra = f"\n\n【补充背景信息】\n{text_info}" if text_info else ""

    prompt = f"""{EXPERT_PERSONA}

请对以下电商详情页截图进行深度分析。

【分析任务 — 详情页表达逻辑与策略】

1. **内容逻辑线**：详情页从上到下的信息编排逻辑是什么？
   - 是什么 → 为什么好 → 为什么选我 → 怎么用 → 怎么买？
   - 遵循了什么说服模型？（AIDA、FABE、痛点→解决方案→背书→促转化 等）

2. **核心说服路径**：详情页是如何一步步说服消费者下单的？
   - 保健品详情页关键说服环节：功效可信度、原料/配方公开、生产资质、临床/检测背书、用户口碑

3. **信任构建手段**：用了哪些信任背书？
   - 常见手段：品牌历史、GMP认证、检测报告、专家推荐、明星代言、销量数据、用户证言

4. **成分与功效表达**：产品核心成分/配方如何呈现？功效宣称的表达方式？
   - 在合规边界内做到了什么程度？

5. **视觉表达风格**：整体设计风格是什么？
   - 品质感/亲民感/专业医疗感/年轻时尚感？与产品定位是否匹配？

6. **差异化表达**：相比同类竞品，详情页突出了什么独特价值？

7. **不足与改进方向**：从转化率角度看，有哪些可以优化的点？
{extra}

【输出格式】请严格输出 JSON：

```json
{{
    "content_logic_flow": "详情页内容逻辑线描述（需要详细，100-200字）",
    "persuasion_path": "核心说服路径分析（100-150字）",
    "trust_building": {{
        "methods": ["信任手段1", "信任手段2", "信任手段3"],
        "evaluation": "信任构建效果评价"
    }},
    "ingredients_expression": {{
        "core_ingredients": "核心成分/配方描述",
        "efficacy_claims": ["功效宣称1", "功效宣称2"],
        "expression_technique": "功效表达技巧分析"
    }},
    "visual_style": {{
        "style_keywords": "设计风格关键词",
        "positioning_match": "与产品定位匹配度评价"
    }},
    "differentiation": "差异化表达分析",
    "weaknesses": "不足之处",
    "improvement_suggestions": ["优化建议1", "优化建议2"]
}}
```"""

    result_text = _call_gemini_vision(image_paths, prompt)
    return _parse_json_response(result_text)


# ============================================================
# PDF 图片分析（图片型 PDF 用视觉 API 识别）
# ============================================================

def analyze_pdf_images(image_paths: list, pdf_filename: str = "", text_info: str = "") -> dict:
    """
    用 Gemini 视觉 API 分析图片型 PDF（如支付人群画像、数据截图等）

    Args:
        image_paths: PDF 转出的图片路径列表
        pdf_filename: 原 PDF 文件名
        text_info: 补充信息

    Returns:
        dict: PDF 内容分析结果
    """
    extra = f"\n\n【补充背景信息】\n{text_info}" if text_info else ""

    prompt = f"""{EXPERT_PERSONA}

以下是从 PDF 文件「{pdf_filename}」中提取的页面图片，请仔细识别图片中的所有文字、数据、图表信息。

【分析任务 — PDF 内容识别与分析】

1. **内容识别**：请完整识别图片中的所有文字内容、数据表格、图表数据
2. **数据提取**：如果包含数据图表（柱状图、饼图、折线图等），请提取关键数据点
3. **信息结构化**：将识别的内容按逻辑组织，方便后续分析使用
4. **专业解读**：从保健品电商运营角度，对识别的内容给出专业解读

请特别注意：
- 如果是人群画像类内容，提取年龄分布、性别比例、地域分布、消费能力等关键维度
- 如果是数据报表类内容，提取核心指标数值和趋势
- 如果是资质/证书类内容，识别关键认证信息
{extra}

【输出格式】请严格输出 JSON：

```json
{{
    "pdf_source": "{pdf_filename}",
    "content_type": "图片中内容的类型描述（如：支付人群画像、数据报表、资质证书等）",
    "extracted_text": "识别出的完整文字内容",
    "key_data": {{
        "数据维度1": "数据值或描述",
        "数据维度2": "数据值或描述"
    }},
    "charts_data": [
        {{
            "chart_type": "图表类型",
            "chart_title": "图表标题",
            "data_points": "关键数据点"
        }}
    ],
    "professional_insights": "从电商运营角度的专业解读（100-200字）",
    "summary": "核心信息摘要（50-100字）"
}}
```"""

    result_text = _call_gemini_vision(image_paths, prompt)
    return _parse_json_response(result_text)


def cleanup_pdf_temp_images(folder_path: str):
    """
    清理 PDF 转出的临时图片文件

    Args:
        folder_path: 竞品文件夹路径
    """
    folder = Path(folder_path)
    for f in folder.rglob("_pdf_page_*.png"):
        try:
            f.unlink()
        except Exception:
            pass


# ============================================================
# 维度三~九：7个数据维度综合分析（纯文本 API，不消耗视觉配额）
# （规格设计、活动促销、评价分析、销售数据、SKU、搜索词、流量）
# ============================================================

def analyze_data_dimensions(data_text: str, text_info: str = "") -> dict:
    """
    基于结构化数据（Excel + PDF 文本），一次性分析 7 个数据维度
    使用纯文本 API，不传图片，避免占用视觉 API 配额

    Args:
        data_text: Excel/PDF 数据转换的文本
        text_info: 补充信息

    Returns:
        dict: 包含 7 个数据维度的分析结果
    """
    extra = f"\n\n【补充背景信息】\n{text_info}" if text_info else ""

    prompt = f"""{EXPERT_PERSONA}

以下是一个保健品竞品的电商运营数据（来自生意参谋等数据工具的导出表格，以及支付人群画像等信息），请基于这些数据进行多维度深度分析。

【原始数据】
{data_text if data_text else "（未提供结构化数据）"}
{extra}

【分析任务 — 7个数据维度深度分析】

请对以下每个维度进行专业分析。数据来源是电商后台的真实运营数据，请仔细阅读每个Sheet的表头和数据内容，提取有价值的信息。如果某维度的数据不完整或缺失，请基于已有数据和行业经验做合理推断，并标注"[推断]"。

**维度1：规格设计分析**
- 产品有哪些SKU规格？（克重/瓶数/盒数等）
- 规格设计的定价策略逻辑？（引流款/利润款/组合装的梯度设计）
- 与保健品行业常见规格设计相比的优劣

**维度2：活动与促销策略**
- 当前参加了哪些平台活动？（直降/满减/买赠/会员价/秒杀/百亿补贴等）
- 优惠券设置策略
- 促销力度在同品类中处于什么水平？

**维度3：评价分析**
- 评价总数量是什么量级？（百级/千级/万级/十万级）
- 好评率/中差评率
- 好评中消费者最常提到的关键词和关注点是什么？
- 差评/中评中消费者主要吐槽什么？
- 从评价中能洞察到什么消费者真实需求？

**维度4：近30日销售数据分析**
- 近30天的销售额/销量数据
- 日均销售水平
- 是否有明显的销售波动？可能原因是什么？
- 在品类中处于什么销售水平？

**维度5：SKU销售数据分析**
- 各规格SKU的销售占比
- 哪个SKU最畅销？为什么？
- SKU之间的价格带分布是否合理？
- 是否有明显的引流款和利润款策略？

**维度6：搜索词与转化分析**
- 主要引流搜索词有哪些？
- 各搜索词的访客量和转化率情况
- 品牌词 vs 品类词 vs 长尾词的分布
- 搜索词策略有什么特点？

**维度7：流量来源分析**
- 各流量渠道（搜索/推荐/活动/直播/站外等）的占比
- 主要流量来源是什么？
- 付费流量和免费流量的比例
- 流量结构是否健康？有什么风险点？

【输出格式】请严格输出 JSON：

```json
{{
    "spec_design": {{
        "sku_list": "各SKU规格与价格描述",
        "pricing_logic": "定价策略逻辑分析",
        "evaluation": "规格设计评价",
        "data_available": true
    }},
    "promotions": {{
        "current_activities": ["活动1", "活动2"],
        "coupon_strategy": "优惠券策略描述",
        "promo_intensity": "促销力度评价（与同品类对比）",
        "data_available": true
    }},
    "reviews_analysis": {{
        "total_volume": "评价数量量级",
        "positive_rate": "好评率",
        "positive_keywords": ["好评关键词1", "好评关键词2", "好评关键词3"],
        "negative_keywords": ["差评关键词1", "差评关键词2"],
        "consumer_insights": "从评价中洞察的消费者真实需求",
        "data_available": true
    }},
    "sales_30days": {{
        "total_sales": "近30天销售额/销量",
        "daily_average": "日均销售",
        "trend_analysis": "销售趋势分析",
        "category_position": "品类中的销售水平定位",
        "data_available": true
    }},
    "sku_sales": {{
        "top_sku": "最畅销SKU及原因",
        "sales_distribution": "各SKU销售占比分析",
        "strategy_insight": "引流款/利润款策略洞察",
        "data_available": true
    }},
    "search_terms": {{
        "top_terms": ["搜索词1", "搜索词2", "搜索词3"],
        "conversion_analysis": "转化率分析",
        "keyword_strategy": "搜索词策略特点",
        "brand_vs_category": "品牌词与品类词比例分析",
        "data_available": true
    }},
    "traffic_sources": {{
        "channel_distribution": "各渠道流量占比描述",
        "primary_source": "主要流量来源",
        "paid_vs_organic": "付费/免费流量比例",
        "health_assessment": "流量结构健康度评估",
        "data_available": true
    }}
}}
```

**重要**：如果某维度数据完全缺失，将 `data_available` 设为 false，其他字段用 "[数据不足，需补充]" 填充。"""

    result_text = _call_gemini_text(prompt)
    return _parse_json_response(result_text)


# ============================================================
# 综合竞品评估总结
# ============================================================

def generate_competitive_summary(all_analysis: dict, our_product_info: str = "") -> dict:
    """
    基于所有维度的分析结果，生成综合竞品评估

    Args:
        all_analysis: 所有维度的分析结果合并
        our_product_info: 我方产品信息（可选）

    Returns:
        dict: 综合评估与策略建议
    """
    our_info = f"\n\n【我方产品信息】\n{our_product_info}" if our_product_info else ""

    prompt = f"""{EXPERT_PERSONA}

以下是对一个保健品竞品的多维度深度分析数据，请生成综合竞品评估报告。

【竞品多维度分析数据】
{json.dumps(all_analysis, ensure_ascii=False, indent=2)}
{our_info}

【综合评估任务】

1. **核心优势总结**：这个竞品最值得警惕的竞争优势是什么？（3-5条）
2. **核心劣势总结**：这个竞品有哪些明显短板可以作为我方突破口？（3-5条）
3. **差异化机会点**：基于竞品分析，我方可以从哪些角度做差异化？
   - 给出3个具体可操作的差异化方向
4. **竞争策略建议**：如果要与这个竞品直接竞争，你的操盘建议是什么？
   - 从产品力、营销力、渠道力三个维度各给1条建议
5. **建议深入分析方向**：基于当前分析，你认为还有哪些方向值得进一步深入研究？

【输出格式】请严格输出 JSON：

```json
{{
    "competitor_strengths": ["优势1", "优势2", "优势3"],
    "competitor_weaknesses": ["劣势1", "劣势2", "劣势3"],
    "differentiation_opportunities": [
        {{
            "direction": "差异化方向",
            "how_to": "具体操作建议"
        }}
    ],
    "competition_strategy": {{
        "product": "产品力策略建议",
        "marketing": "营销力策略建议",
        "channel": "渠道力策略建议"
    }},
    "further_analysis_suggestions": ["建议深入方向1", "建议深入方向2"],
    "overall_threat_level": "竞争威胁评级（高/中/低）及理由"
}}
```"""

    result_text = _call_gemini_text(prompt)
    return _parse_json_response(result_text)


# ============================================================
# 批量分析入口
# ============================================================

def analyze_competitor(competitor_data: dict, delay: float = 1.0, our_product_info: str = "") -> dict:
    """
    对单个竞品执行全维度分析

    Args:
        competitor_data: image_reader 返回的竞品数据
        delay: API 调用间隔（秒）
        our_product_info: 我方产品信息

    Returns:
        dict: 合并后的完整分析结果
    """
    categories = competitor_data["categories"]
    data_files = competitor_data.get("data_files", {})
    text_info = competitor_data.get("text_info", "")

    all_results = {
        "competitor_name": competitor_data["name"],
    }

    # ====== Phase 1: 读取结构化数据 ======
    excel_text = ""
    pdf_image_paths = []  # 图片型 PDF 转出的图片
    if any(data_files.values()):
        print(f"    📊 读取结构化数据...")
        excel_text, pdf_image_paths = read_all_data_files(data_files)
        if excel_text:
            all_results["_has_excel_data"] = True
            print(f"       ✓ 已读取 {len(excel_text)} 字符的文本数据")
        if pdf_image_paths:
            print(f"       ✓ 发现 {len(pdf_image_paths)} 张图片型PDF页面，待视觉分析")

    # ====== Phase 2: 主图卖点与视觉策略 ======
    main_images = categories.get("main_images", [])
    if main_images:
        print(f"    🎨 分析主图卖点与策略（{len(main_images)} 张）...")
        try:
            batch = main_images[:10]
            result = analyze_main_images_strategy(batch, text_info)
            all_results["main_images_strategy"] = result
        except Exception as e:
            print(f"    ⚠ 主图分析失败: {e}")
            all_results["main_images_strategy"] = {"error": str(e)}
        time.sleep(delay)

    # ====== Phase 3: 详情页表达逻辑 ======
    detail_pages = categories.get("detail_pages", [])
    if detail_pages:
        print(f"    📋 分析详情页表达逻辑（{len(detail_pages)} 张）...")
        try:
            batch = detail_pages[:10]
            result = analyze_detail_page_logic(batch, text_info)
            all_results["detail_page_logic"] = result
        except Exception as e:
            print(f"    ⚠ 详情页分析失败: {e}")
            all_results["detail_page_logic"] = {"error": str(e)}
        time.sleep(delay)

    # ====== Phase 3.5: 图片型 PDF 视觉分析 ======
    if pdf_image_paths:
        print(f"    📄 用视觉API分析图片型PDF（{len(pdf_image_paths)} 页）...")
        try:
            # 按每个PDF源文件分组分析
            pdf_groups = {}
            for img_path in pdf_image_paths:
                # 从文件名 _pdf_page_{stem}_{page}.png 提取源文件名
                fname = Path(img_path).stem  # _pdf_page_支付人群画像_1
                parts = fname.replace("_pdf_page_", "").rsplit("_", 1)
                pdf_name = parts[0] if len(parts) > 1 else fname
                if pdf_name not in pdf_groups:
                    pdf_groups[pdf_name] = []
                pdf_groups[pdf_name].append(img_path)

            pdf_analyses = []
            for pdf_name, pages in pdf_groups.items():
                print(f"       🔍 分析: {pdf_name}（{len(pages)} 页）")
                pdf_result = analyze_pdf_images(pages, f"{pdf_name}.pdf", text_info)
                pdf_analyses.append(pdf_result)
                if len(pdf_groups) > 1:
                    time.sleep(delay)

            all_results["pdf_analysis"] = pdf_analyses
            print(f"       ✓ 图片型PDF分析完成")

            # 如果 PDF 分析提取到了文字数据，追加到 excel_text 中供后续数据维度分析
            for pa in pdf_analyses:
                if isinstance(pa, dict) and not pa.get("_parse_failed"):
                    extracted = pa.get("extracted_text", "")
                    summary = pa.get("summary", "")
                    pdf_src = pa.get("pdf_source", "PDF")
                    if extracted or summary:
                        extra_text = f"\n{'#'*50}\n数据来源（PDF视觉识别）: {pdf_src}\n{'#'*50}\n"
                        if extracted:
                            extra_text += f"{extracted}\n"
                        if summary:
                            extra_text += f"摘要: {summary}\n"
                        excel_text = (excel_text + extra_text) if excel_text else extra_text

        except Exception as e:
            print(f"    ⚠ 图片型PDF分析失败: {e}")
            all_results["pdf_analysis"] = [{"error": str(e)}]
        time.sleep(delay)

    # ====== Phase 4: 7个数据维度分析（纯文本API，1次调用） ======
    if excel_text:
        print(f"    📈 分析7个数据维度（规格/活动/评价/销售/SKU/搜索词/流量）...")
        try:
            data_result = analyze_data_dimensions(excel_text, text_info)
            if not data_result.get("_parse_failed"):
                all_results["data_analysis"] = data_result
                print(f"       ✓ 7个数据维度分析完成")
            else:
                print(f"    ⚠ 数据维度分析结果解析失败")
                all_results["data_analysis"] = {}
        except Exception as e:
            print(f"    ⚠ 数据维度分析失败: {e}")
            all_results["data_analysis"] = {"error": str(e)}
        time.sleep(delay)
    else:
        all_results["data_analysis"] = {}

    # ====== Phase 5: 综合竞品评估 ======
    print(f"    🏁 生成综合竞品评估...")
    try:
        summary = generate_competitive_summary(all_results, our_product_info)
        all_results["competitive_summary"] = summary
    except Exception as e:
        print(f"    ⚠ 综合评估生成失败: {e}")
        all_results["competitive_summary"] = {"error": str(e)}

    # ====== 清理临时文件 ======
    if pdf_image_paths:
        comp_path = competitor_data.get("path", "")
        if comp_path:
            cleanup_pdf_temp_images(comp_path)
            print(f"    🧹 已清理PDF临时图片")

    return all_results
